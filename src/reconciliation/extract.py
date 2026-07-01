"""Extractores de los Excel de ALTAS (lo pagado a técnicos) y ANEXOS (lo certificado).

Los archivos reales son irregulares: cabeceras desplazadas, fechas en celdas combinadas,
órdenes ORANGE leídas como float, y add-ons MASMOVIL con sufijo (_AGILETV). Estos extractores
toleran todo eso para que el cruce por número de orden sea fiable.
"""
import re
from datetime import datetime, date
from openpyxl import load_workbook

# Filas de pie/resumen que NO son órdenes reales en las hojas de ALTAS
_ALTAS_STOP = {
    "festivos", "descuentos", "altas en garantia", "altas en garantía", "total", "totales",
    "resumen", "observaciones", "nota", "notas", "subtotal", "irpf", "embargo", "multa",
    "gasolina", "reutilizadas garantia", "averias en garantia", "alta en garantia",
}


def norm_order(value) -> str:
    """Normaliza un nº de orden: mayúsculas, sin espacios, sin el .0 de floats."""
    text = str(value).strip().upper()
    if re.fullmatch(r"\d+\.0+", text):           # 8571935.0 leído como float
        text = text.split(".")[0]
    return text


def base_order(order: str) -> str:
    """Quita el sufijo add-on tipo _AGILETV (último segmento sin dígitos)."""
    if "_" in order:
        head, _, tail = order.rpartition("_")
        if head and not re.search(r"\d", tail):
            return head
    return order


def order_key(value) -> str:
    return base_order(norm_order(value))


def linea_of(value) -> str:
    """Clasifica la línea de negocio por el formato del número de orden."""
    order = norm_order(value)
    if order.startswith("SC2026"):
        return "ALARMAS"
    if order.startswith(("MYSIM_", "WD_", "ATC-", "PP")):
        return "MASMOVIL"
    return "ORANGE"


def _is_order(value) -> bool:
    text = str(value).strip()
    if not text or text.lower() in _ALTAS_STOP:
        return False
    return bool(re.search(r"\d", text)) and len(text) >= 5


def parse_date(value):
    """Parsea fechas en los múltiples formatos que aparecen (D/M, M/D, ISO, datetime)."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if match:
        first, second, year = (int(match.group(i)) for i in (1, 2, 3))
        # Prioriza D/M salvo que sea imposible (primer campo > 12 => es día)
        day, month = (first, second) if first > 12 else (second, first) if second > 12 else (first, second)
        try:
            return date(year, month, day)
        except ValueError:
            return None
    return None


def extract_altas(path: str) -> list[dict]:
    """Lee un Excel de ALTAS (una hoja por técnico) y devuelve una fila por orden pagada.

    Cada dict: {tecnico, orden, codigo, fecha (date|None), precio (float|None)}.
    Rellena hacia abajo la fecha de las celdas combinadas y captura el precio pagado
    al técnico (4ª columna 'TECNICO', que contiene el importe en euros).
    """
    workbook = load_workbook(path, data_only=True)
    altas = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = col_orden = col_codigo = col_fecha = col_precio = None
        for idx, row in enumerate(rows):
            cells = [str(c).strip().upper() if c is not None else "" for c in row]
            if "ORDEN" in cells and "CODIGO" in cells:
                header_idx = idx
                col_orden = cells.index("ORDEN")
                col_codigo = cells.index("CODIGO")
                col_fecha = cells.index("FECHA") if "FECHA" in cells else None
                col_precio = cells.index("TECNICO") if "TECNICO" in cells else None
                break
        if header_idx is None:
            continue

        last_date = None
        for row in rows[header_idx + 1:]:
            if col_fecha is not None and col_fecha < len(row) and row[col_fecha] is not None:
                parsed = parse_date(row[col_fecha])
                if parsed:
                    last_date = parsed
            orden = str(row[col_orden]).strip() if col_orden < len(row) and row[col_orden] is not None else ""
            if not _is_order(orden):
                continue
            codigo = ""
            if col_codigo is not None and col_codigo < len(row) and row[col_codigo] is not None:
                codigo = str(row[col_codigo]).strip()
            precio = None
            if col_precio is not None and col_precio < len(row) and row[col_precio] is not None:
                try:
                    precio = float(row[col_precio])
                except (TypeError, ValueError):
                    precio = None
            altas.append({
                "tecnico": sheet.title.strip(),
                "orden": orden,
                "codigo": codigo,
                "fecha": last_date,
                "precio": precio,
            })
    return altas


def extract_anexo(path: str) -> list[dict]:
    """Lee un anexo de certificación y devuelve una fila por orden certificada.

    Cada dict: {orden, importe (float|None), fecha (date|None), tecnico, tipo}.
    Detecta la tabla por la cabecera que contiene 'IMPORTE TOTAL'; la columna de orden
    es la primera con 'ORDEN', 'ITEM' o 'PRESUPUESTO'.
    """
    workbook = load_workbook(path, data_only=True)
    lineas = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = None
        cols = {}
        for idx, row in enumerate(rows):
            cells = [str(c).strip().upper() if c is not None else "" for c in row]
            if any("IMPORTE TOTAL" in c for c in cells):
                header_idx = idx
                for j, cell in enumerate(cells):
                    if "IMPORTE TOTAL" in cell:
                        cols["importe"] = j
                    if ("ORDEN" in cell or "ITEM" in cell or "PRESUPUESTO" in cell) and "orden" not in cols:
                        cols["orden"] = j
                    if "TECNICO" in cell:
                        cols["tecnico"] = j
                    if "TIPO ORDEN" in cell:
                        cols["tipo"] = j
                    if cell == "FECHA":
                        cols["fecha"] = j
                break
        if header_idx is None or "orden" not in cols:
            continue

        for row in rows[header_idx + 1:]:
            orden = str(row[cols["orden"]]).strip() if cols["orden"] < len(row) and row[cols["orden"]] is not None else ""
            if not _is_order(orden):
                continue
            importe = None
            if "importe" in cols and cols["importe"] < len(row) and row[cols["importe"]] is not None:
                try:
                    importe = float(row[cols["importe"]])
                except (TypeError, ValueError):
                    importe = None
            fecha = parse_date(row[cols["fecha"]]) if "fecha" in cols and cols["fecha"] < len(row) else None
            tecnico = str(row[cols["tecnico"]]).strip() if "tecnico" in cols and cols["tecnico"] < len(row) and row[cols["tecnico"]] is not None else ""
            tipo = str(row[cols["tipo"]]).strip() if "tipo" in cols and cols["tipo"] < len(row) and row[cols["tipo"]] is not None else ""
            lineas.append({
                "orden": orden,
                "importe": importe,
                "fecha": fecha,
                "tecnico": tecnico,
                "tipo": tipo,
            })
    return lineas
