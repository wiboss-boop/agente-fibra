#!/usr/bin/env python3
"""Genera ALTAS_<MES>_<AÑO>.xlsx desde el Sheet mensual del agente.

Reglas acordadas con el usuario:
- Todo es del mes en curso. El DÍA de la fecha es correcto; mes/año se fuerzan.
- Precio al técnico: se toma de la columna TECNICO; si está vacío se rellena por código.
- Se excluyen filas que no son órdenes (SIN ALTAS, pies de nómina, festivos, etc.).
- NO se puede pagar una orden dos veces:
    · Mismo orden + mismo día en un técnico → duplicado técnico (el agente re-apila su
      volcado diario debajo del bloque ya cerrado). Se colapsa a una sola fila,
      prefiriendo la que trae precio del contador.
    · Mismo orden en días distintos (mismo técnico) o en dos técnicos → AMBIGUO: no se
      resuelve solo; se deja en el archivo y se REPORTA para decisión manual.
- Formato de salida = igual que ALTAS_MAYO_2026.xlsx: una hoja por técnico,
  fila 1 título 'NOMBRE — <MES> <AÑO>', fila 2 cabecera FECHA|ORDEN|CODIGO|TECNICO.
- Cada hoja lleva un pie con SUBTOTAL ALTAS, el detalle de la pestaña 'Descuentos' del
  Sheet (adelantos, gasolina, multas…), TOTAL DESCUENTOS y TOTAL A PAGAR, con fórmulas
  para poder retocarlo a mano. Sustituye al pie que antes se tecleaba en cada hoja.

Uso:  gen_altas_mensual.py --mes JULIO [--anio 2026] [--sheet ID] [--out X] [--write]
"""
import argparse
import re
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.reconciliation.extract import norm_order
from src.sheets.auth import get_sheets_service

# Sheet de origen por mes (el ID cambia cada mes; ver ACTIVE_SHEET_ID en Railway)
SHEETS = {
    ("JUNIO", 2026): "1JjUY08AJmICiPmc9xXDJEdqKFIv2rIln6AYYWUxzMwI",
    ("JULIO", 2026): "1Suv1VbJ232NA7PtsUaxwA_-QwXquz8mRbcv3qvTXe6A",
    ("AGOSTO", 2026): "1-GRoSJIK1X6dP6AyG0ax_lsfMcPqSjfM9oqwNdilbtc",
}
MESES = {"ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
         "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10,
         "NOVIEMBRE": 11, "DICIEMBRE": 12}
OUT_DIR = ("/Users/samaro/Library/CloudStorage/GoogleDrive-salamanca118@gmail.com/"
           "Mi unidad/SECOMCOL/CONTABILIDAD/ALTAS POR TECNICO/2026")
TECS = ["CRISTIAN", "MARTIN", "JAMES", "JEAN", "YOHAN", "ERCS",
        "HANS", "JOEL", "DIANA", "AYMAN", "LUIS E"]

STOP = {"FESTIVOS", "DESCUENTOS", "TOTAL", "TOTALES", "SUBTOTAL", "IRPF",
        "SECOMCOL", "SEG SOCIAL", "G BRUTA", "SIN ALTAS", "SIN PARTE",
        "RESUMEN", "OBSERVACIONES", "EMBARGO", "MULTA", "GASOLINA"}

# Precio al técnico por código (derivado de junio; consistente con mayo)
PRECIO_TECNICO = {
    "AVERIA OK": 10, "MM01": 27, "MM02": 27, "MM03": 27, "MM04": 27,
    "MM05": 27, "MM06": 27, "MM12": 27, "MM17": 17, "ZA_DESMONTAJE": 10,
    "ZA_INC/MTO/AMP": 14, "ZA_INCIDENCIAS": 14, "ZA_INSTALACION": 30,
    "ZA_TRASLADO": 40,
}

# ---------------------------------------------------------------------------
# Resoluciones manuales de duplicados ambiguos (decisión del usuario).
# SON ESPECÍFICAS DE CADA MES: la misma orden no puede pagarse dos veces, pero la
# decisión de dónde pagarla la toma el usuario mes a mes. NUNCA heredar las de otro mes.
# ---------------------------------------------------------------------------
RESOLUCIONES = {
    ("JUNIO", 2026): {
        # (técnico, clave) -> día a conservar; se eliminan las demás fechas de esa orden.
        "keep_one_day": {
            ("JOEL", "OT107152"): 22,   # instalación pagada 3 veces (22/25/26) -> una sola
        },
        # (técnico, clave) -> nota a escribir en la fila (col NOTA).
        "notes": {
            ("JEAN", "SC2026202530"): "Tambien reportada por MARTIN 18/06; se conserva aqui por reporte previo (17/06)",
            ("MARTIN", "SC2026202530"): "Duplicada con JEAN 17/06 — NO se paga aqui",
        },
        # (técnico, clave) que se dejan VISIBLES pero sin pago (precio 0).
        "no_pay": {
            ("MARTIN", "SC2026202530"),
        },
    },
}
KEEP_ONE_DAY, NOTES, NO_PAY = {}, {}, set()   # se rellenan en main() según el mes


def is_order(orden) -> bool:
    text = str(orden).strip()
    return bool(text) and text != "-" and text.upper() not in STOP \
        and bool(re.search(r"\d", text)) and len(text) >= 5


def day_of(fecha):
    match = re.match(r"(\d{1,2})", str(fecha).strip())
    return int(match.group(1)) if match else None


def money(value):
    text = str(value).replace("€", "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def read_altas(service):
    """Lee todas las hojas de técnico y devuelve dicts normalizados (una por fila válida)."""
    altas = []
    stats = defaultdict(lambda: {"fixed_date": 0, "filled_price": 0, "no_price": 0})
    for tab in TECS:
        rows = service.spreadsheets().values().get(
            spreadsheetId=SID, range=f"'{tab}'!A3:E"
        ).execute().get("values", [])
        for row in rows:
            row = row + [""] * (5 - len(row))
            fecha, orden, codigo, _precio_contratista, tecnico = row
            if not is_order(orden):
                continue
            day = day_of(fecha)
            if day is None or not (1 <= day <= 31):
                print(f"  ⚠ {tab}: fecha ilegible {fecha!r} orden={orden} — omitida")
                continue
            fecha_ok = f"{day:02d}/{MES_NUM:02d}/{ANIO}"
            if str(fecha).strip().replace(" ", "") != fecha_ok:
                stats[tab]["fixed_date"] += 1
            codigo = str(codigo).strip().upper()
            precio = money(tecnico)
            had_price = precio is not None
            if not had_price:
                precio = PRECIO_TECNICO.get(codigo)
                if precio is not None:
                    stats[tab]["filled_price"] += 1
                else:
                    stats[tab]["no_price"] += 1
            altas.append({
                "tecnico": tab, "orden": str(orden).strip(),
                "clave": norm_order(orden), "codigo": codigo,
                "dia": day, "fecha": fecha_ok, "precio": precio,
                "had_price": had_price, "nota": "",
            })
    return altas, stats


def read_descuentos(service):
    """Lee la pestaña Descuentos del Sheet -> {tecnico: [(fecha, concepto, monto)]}."""
    rows = service.spreadsheets().values().get(
        spreadsheetId=SID, range="'Descuentos'!A2:D"
    ).execute().get("values", [])
    out = defaultdict(list)
    for row in rows:
        row = row + [""] * (4 - len(row))
        fecha, tecnico, concepto, monto = (str(x).strip() for x in row[:4])
        importe = money(monto)
        if not tecnico or importe is None:
            if any((fecha, tecnico, concepto, monto)):
                print(f"  ⚠ Descuentos: fila ignorada {row!r}")
            continue
        out[tecnico.upper()].append((fecha, concepto, importe))
    return out


def escribir_pie(sheet, fila_ultima_alta, descuentos, header_font):
    """Pie por técnico: subtotal de altas, detalle de descuentos y total a pagar."""
    fila_sub = fila_ultima_alta + 2
    sheet.cell(fila_sub, 2, "SUBTOTAL ALTAS").font = header_font
    if fila_ultima_alta >= 3:
        sheet.cell(fila_sub, 4, f"=SUM(D3:D{fila_ultima_alta})")
    else:
        sheet.cell(fila_sub, 4, 0)

    fila = fila_sub + 2
    sheet.cell(fila, 2, "DESCUENTOS").font = header_font
    primera = fila + 1
    for i, (fecha, concepto, importe) in enumerate(descuentos):
        sheet.cell(primera + i, 1, fecha)
        sheet.cell(primera + i, 2, concepto)
        sheet.cell(primera + i, 4, importe)
    ultima = primera + len(descuentos) - 1

    fila_tot_desc = (ultima if descuentos else fila) + 1
    sheet.cell(fila_tot_desc, 2, "TOTAL DESCUENTOS").font = header_font
    sheet.cell(fila_tot_desc, 4,
               f"=SUM(D{primera}:D{ultima})" if descuentos else 0)

    fila_pagar = fila_tot_desc + 2
    sheet.cell(fila_pagar, 2, "TOTAL A PAGAR").font = header_font
    cell = sheet.cell(fila_pagar, 4, f"=D{fila_sub}-D{fila_tot_desc}")
    cell.font = header_font


def apply_resolutions(altas):
    """Aplica las decisiones manuales sobre duplicados ambiguos. Devuelve las altas finales."""
    result = []
    for a in altas:
        sig = (a["tecnico"], a["clave"])
        keep_day = KEEP_ONE_DAY.get(sig)
        if keep_day is not None and a["dia"] != keep_day:
            continue  # orden repetida en otro día -> se descarta
        if keep_day is not None and a["dia"] == keep_day:
            a["nota"] = f"Orden única (registrada varias veces en el mes; se paga 1 vez)"
        if sig in NO_PAY:
            a["precio"] = 0
        if sig in NOTES:
            a["nota"] = NOTES[sig]
        result.append(a)
    return result


def dedup_same_day(altas):
    """Colapsa (técnico, orden, día) repetidos a una sola fila; prefiere la que trae precio.

    Devuelve (altas_unicas, removidas_por_tecnico).
    """
    best = {}
    removed = defaultdict(int)
    order_seen = defaultdict(int)
    for a in altas:
        key = (a["tecnico"], a["clave"], a["dia"])
        if key not in best:
            best[key] = a
            order_seen[key] += 1
        else:
            order_seen[key] += 1
            # preferir la fila con precio de fuente (contador)
            if a["had_price"] and not best[key]["had_price"]:
                best[key] = a
            removed[a["tecnico"]] += 1
    return list(best.values()), removed


def find_flags(altas):
    """Duplicados que NO se resuelven solos: mismo orden en días distintos o en 2 técnicos."""
    by_tec_order = defaultdict(list)     # (tec, clave) -> [dias]
    by_order = defaultdict(set)          # clave -> {tecnicos}
    detail = defaultdict(list)           # clave -> [(tec, fecha, precio)]
    for a in altas:
        by_tec_order[(a["tecnico"], a["clave"])].append(a["dia"])
        by_order[a["clave"]].add(a["tecnico"])
        detail[a["clave"]].append((a["tecnico"], a["fecha"], a["precio"]))
    multidia = {k: v for k, v in by_tec_order.items() if len(set(v)) > 1}
    multitec = {k: v for k, v in by_order.items() if len(v) > 1}
    return multidia, multitec, detail


def main() -> None:
    global SID, MES, MES_NUM, ANIO, OUT, KEEP_ONE_DAY, NOTES, NO_PAY

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--mes", required=True, help="mes en mayúsculas, p.ej. JULIO")
    ap.add_argument("--anio", type=int, default=2026)
    ap.add_argument("--sheet", help="ID del Sheet origen (por defecto, el de SHEETS)")
    ap.add_argument("--out", help="ruta del .xlsx (por defecto ALTAS_<MES>_<AÑO>.xlsx)")
    ap.add_argument("--write", action="store_true", help="guarda el .xlsx (si no, dry-run)")
    args = ap.parse_args()

    MES = args.mes.upper()
    ANIO = args.anio
    if MES not in MESES:
        ap.error(f"mes desconocido: {MES}")
    MES_NUM = MESES[MES]
    SID = args.sheet or SHEETS.get((MES, ANIO))
    if not SID:
        ap.error(f"no hay Sheet conocido para {MES} {ANIO}; pásalo con --sheet")
    OUT = args.out or f"{OUT_DIR}/ALTAS_{MES}_{ANIO}.xlsx"

    res = RESOLUCIONES.get((MES, ANIO), {})
    KEEP_ONE_DAY = res.get("keep_one_day", {})
    NOTES = res.get("notes", {})
    NO_PAY = res.get("no_pay", set())
    if not res:
        print(f"ℹ Sin resoluciones manuales registradas para {MES} {ANIO}: "
              f"los duplicados ambiguos saldrán todos reportados abajo.")

    write = args.write
    service = get_sheets_service()

    altas, stats = read_altas(service)
    descuentos = read_descuentos(service)
    altas, removed = dedup_same_day(altas)
    altas = apply_resolutions(altas)
    multidia, multitec, detail = find_flags(altas)

    workbook = Workbook()
    workbook.remove(workbook.active)
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    per_tec = defaultdict(list)
    for a in altas:
        per_tec[a["tecnico"]].append(a)

    for tab in TECS:
        items = sorted(per_tec[tab], key=lambda a: (a["dia"], a["orden"]))
        sheet = workbook.create_sheet(tab)
        sheet["A1"] = f"{tab} — {MES} {ANIO}"
        sheet["A1"].font = title_font
        for col, name in zip("ABCDE", ["FECHA", "ORDEN", "CODIGO", "TECNICO", "NOTA"]):
            cell = sheet[f"{col}2"]
            cell.value = name
            cell.font = header_font
            cell.fill = header_fill
        for i, a in enumerate(items, start=3):
            for j, val in enumerate([a["fecha"], a["orden"], a["codigo"], a["precio"], a["nota"]]):
                sheet.cell(row=i, column=j + 1, value=val)
        for col, width in zip("ABCDE", (12, 22, 16, 10, 52)):
            sheet.column_dimensions[col].width = width
        escribir_pie(sheet, 2 + len(items), descuentos.get(tab, []), header_font)

    print("\nTécnico    altas  fecha_corr  precio_relleno  dup_mismodia_elim  sin_precio")
    for tab in TECS:
        s = stats[tab]
        print(f"  {tab:8} {len(per_tec[tab]):5}   {s['fixed_date']:8}   "
              f"{s['filled_price']:12}   {removed[tab]:15}   {s['no_price']}")
    print(f"  {'TOTAL':8} {len(altas):5}   {sum(s['fixed_date'] for s in stats.values()):8}   "
          f"{sum(s['filled_price'] for s in stats.values()):12}   "
          f"{sum(removed.values()):15}   {sum(s['no_price'] for s in stats.values())}")

    print("\nDescuentos aplicados en el pie de cada hoja:")
    for tab in TECS:
        d = descuentos.get(tab, [])
        if d:
            print(f"  {tab:9} {len(d):3} apuntes   {sum(x[2] for x in d):9.2f} €")
    huerfanos = set(descuentos) - set(TECS)
    if huerfanos:
        print(f"  ⚠ descuentos de técnicos sin hoja (NO aplicados): {sorted(huerfanos)}")

    print("\n⚠ REVISAR — mismo orden en DÍAS distintos (mismo técnico):")
    if multidia:
        for (tec, clave), dias in multidia.items():
            print(f"  {tec}: {clave} en días {sorted(set(dias))}")
    else:
        print("  ninguno")

    print("\n⚠ REVISAR — mismo orden en DOS técnicos:")
    if multitec:
        for clave, tecs in multitec.items():
            print(f"  {clave}: {detail[clave]}")
    else:
        print("  ninguno")

    if write:
        workbook.save(OUT)
        print(f"\n✅ Escrito: {OUT}")
    else:
        print("\n(dry-run — usa --write para guardar el .xlsx)")


if __name__ == "__main__":
    main()
