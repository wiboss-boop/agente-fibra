#!/usr/bin/env python3
"""Genera REGISTRO_JORNADA_<MES>.xlsx a partir del registro del mes anterior.

- Solo se llenan los días que el técnico trabajó (= días con al menos un alta en
  ALTAS_<MES>_<AÑO>.xlsx). Los demás quedan en blanco.
- Cada hoja conserva encabezado, formato, fórmula de total y firmas.
- Patrón por día:
    FULL (7h): 09:00 / 16:00 · 13;00 / 19:00 · ordinarias 7 · total 7
    PART (2h): 11:00 · 13:00 · ordinarias 2 · total 2      (JOEL, AYMAN)
- DIEGO ya no trabaja → hoja en blanco.
- OJO: las hojas NO están todas alineadas — algunas no tienen fila "Periodo de
  liquidación", así que su rejilla arranca una fila más arriba. Por eso la rejilla se
  localiza por contenido ("Día del mes" / "Total horas mes"), no por fila fija.
- OJO 2: la plantilla es el registro del mes anterior, al que este mismo script le borró
  del calendario los días que sobraban. Al pasar de un mes de 30 días a uno de 31 hay que
  RESTAURAR el día 31 en la col A (la fila sigue existiendo y la fórmula del total ya la
  cubre); si no, ese día no se registra.

Uso:  gen_jornada_mensual.py --mes JULIO [--anio 2026] [--plantilla X] [--altas Y] [--out Z]
"""
import argparse
import calendar
import os
import sys
from datetime import date, time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from src import meses
from src.reconciliation.extract import parse_date

BASE = os.environ.get(
    "SECOMCOL_BASE",
    "/Users/samaro/Library/CloudStorage/GoogleDrive-salamanca118@gmail.com/"
    "Mi unidad/SECOMCOL",
)
CONTABILIDAD = f"{BASE}/CONTABILIDAD"

MESES = {nombre: i + 1 for i, nombre in enumerate(meses.MESES)}


def altas_dir(anio: int) -> str:
    return f"{CONTABILIDAD}/ALTAS POR TECNICO/{anio}"


def jornada_dir(anio: int) -> str:
    """Los registros se archivan por año; el de DICIEMBRE vive en la carpeta del año anterior."""
    return f"{CONTABILIDAD}/registro jornada/{anio}"

# hoja del registro -> (técnico en ALTAS, patrón)   None = dejar en blanco
# La hoja de JAMES se llamaba ALVARO hasta que el usuario la renombró en junio;
# se resuelve contra los nombres reales de la plantilla.
MAPEO = {
    "CRISTIAN":   ("CRISTIAN", "FULL"),
    "HANS":       ("HANS",     "FULL"),
    "JEAN MARCO": ("JEAN",     "FULL"),
    "JOEL":       ("JOEL",     "PART"),
    "JAMES":      ("JAMES",    "FULL"),
    "ERCS":       ("ERCS",     "FULL"),
    "YOHAN":      ("YOHAN",    "FULL"),
    "MARTIN":     ("MARTIN",   "FULL"),
    "DIANA":      ("DIANA",    "FULL"),
    "DIEGO":      (None,       "FULL"),   # ya no trabaja; hoja en blanco
    "AYMAN":      ("AYMAN",    "PART"),
}
ALIAS_HOJA = {"JAMES": "ALVARO"}   # nombre alternativo si la plantilla es antigua


def dias_trabajados(path, mes_num):
    """Devuelve {tecnico: set(días)} a partir de las altas del mes."""
    wb = load_workbook(path, data_only=True)
    dias = {}
    for ws in wb.worksheets:
        s = set()
        for row in list(ws.iter_rows(values_only=True))[2:]:
            # el pie (SUBTOTAL ALTAS / DESCUENTOS / TOTAL…) no son días trabajados:
            # sus filas también traen fecha y concepto, así que hay que parar aquí.
            if row and row[1] == "SUBTOTAL ALTAS":
                break
            if not row or row[0] is None or row[1] in (None, "", "ORDEN"):
                continue
            fecha = parse_date(row[0])
            if fecha and fecha.month == mes_num:
                s.add(fecha.day)
        dias[ws.title] = s
    return dias


def _set(ws, fila, col, valor, fmt=None):
    cell = ws.cell(row=fila, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = valor
    if fmt:
        cell.number_format = fmt


def _texto(ws, fila):
    v = ws.cell(fila, 1).value
    return str(v).strip() if v is not None else ""


def localizar(ws):
    """Devuelve (fila_dia1, fila_total, fila_periodo|None, fila_recibido|None)."""
    fila_header = fila_total = fila_periodo = fila_recibido = None
    for r in range(1, ws.max_row + 1):
        t = _texto(ws, r).lower()
        if t.startswith("día del mes") or t.startswith("dia del mes"):
            fila_header = r
        elif t.startswith("total horas mes"):
            fila_total = r
        elif t.startswith("periodo de liquidaci"):
            fila_periodo = r
        elif t.startswith("recibido por el trabajador"):
            fila_recibido = r + 1   # la fecha va en la fila siguiente, col A
    return fila_header + 2, fila_total, fila_periodo, fila_recibido


def escribir_full(ws, fila):
    _set(ws, fila, 2, "09:00")
    _set(ws, fila, 3, time(16, 0), "hh:mm")
    _set(ws, fila, 4, "13;00")
    _set(ws, fila, 5, "19:00")
    _set(ws, fila, 6, 7)
    _set(ws, fila, 9, 7)


def escribir_part(ws, fila):
    _set(ws, fila, 2, time(11, 0), "hh:mm")
    _set(ws, fila, 3, None)
    _set(ws, fila, 4, time(13, 0), "hh:mm")
    _set(ws, fila, 5, None)
    _set(ws, fila, 6, 2)
    _set(ws, fila, 9, 2)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--mes", required=True, help="mes en mayúsculas, p.ej. JULIO")
    ap.add_argument("--anio", type=int, default=2026)
    ap.add_argument("--plantilla", help="xlsx base (por defecto, el registro del mes anterior)")
    ap.add_argument("--altas", help="xlsx de altas (por defecto ALTAS_<MES>_<AÑO>.xlsx)")
    ap.add_argument("--out", help="xlsx de salida (por defecto REGISTRO_JORNADA_<MES>.xlsx)")
    args = ap.parse_args()

    mes = args.mes.upper()
    if mes not in MESES:
        ap.error(f"mes desconocido: {mes}")
    mes_num, anio = MESES[mes], args.anio
    dias_mes = calendar.monthrange(anio, mes_num)[1]
    inicio, fin = date(anio, mes_num, 1), date(anio, mes_num, dias_mes)

    mes_prev_num, anio_prev = meses.anterior(mes_num, anio)
    mes_prev = meses.nombre(mes_prev_num)

    altas_path = args.altas or f"{altas_dir(anio)}/ALTAS_{mes}_{anio}.xlsx"
    out_path = args.out or f"{jornada_dir(anio)}/REGISTRO_JORNADA_{mes}.xlsx"
    # La plantilla es el registro del mes anterior — que en enero está en el año anterior.
    template = args.plantilla or f"{jornada_dir(anio_prev)}/REGISTRO_JORNADA_{mes_prev}.xlsx"

    print(f"Mes:       {mes} {anio} ({dias_mes} días)")
    print(f"Altas:     {altas_path}")
    print(f"Plantilla: {template}")
    print(f"Salida:    {out_path}\n")

    dias = dias_trabajados(altas_path, mes_num)
    wb = load_workbook(template)

    resumen = []
    for hoja, (tecnico, patron) in MAPEO.items():
        if hoja not in wb.sheetnames and ALIAS_HOJA.get(hoja) in wb.sheetnames:
            hoja = ALIAS_HOJA[hoja]      # plantilla antigua: JAMES aún se llama ALVARO
        ws = wb[hoja]
        fila1, fila_total, fila_periodo, fila_recibido = localizar(ws)

        # mapa día -> fila (primera aparición del número de día en col A)
        dia_a_fila = {}
        for r in range(fila1, fila_total):
            v = ws.cell(r, 1).value
            if isinstance(v, int) and 1 <= v <= 31 and v not in dia_a_fila:
                dia_a_fila[v] = r

        # limpiar toda la rejilla (cols B..I) y borrar días que sobran del mes
        for r in range(fila1, fila_total):
            for c in range(2, 10):
                _set(ws, r, c, None)
            v = ws.cell(r, 1).value
            if isinstance(v, int) and v > dias_mes:
                _set(ws, r, 1, None)

        # restaurar días que falten (la plantilla viene de un mes más corto:
        # el generador anterior borró el 31 de la col A, pero la fila sigue ahí)
        faltantes = []
        d, r = max(dia_a_fila, default=0), dia_a_fila.get(max(dia_a_fila, default=0), fila1 - 1)
        while d < dias_mes and r + 1 < fila_total:
            d, r = d + 1, r + 1
            _set(ws, r, 1, d)
            dia_a_fila[d] = r
            faltantes.append(d)
        if faltantes:
            print(f"  ℹ {hoja}: restaurados días {faltantes} en el calendario")
        if max(dia_a_fila, default=0) < dias_mes:
            print(f"  ⚠ {hoja}: la rejilla no tiene sitio para el día {dias_mes}")

        # periodo (si la hoja tiene esa fila) y fecha de recibido
        if fila_periodo:
            _set(ws, fila_periodo, 4, inicio)
            _set(ws, fila_periodo, 10, fin)
        if fila_recibido:
            _set(ws, fila_recibido, 1, fin)

        trabajados = sorted(d for d in dias.get(tecnico, set()) if 1 <= d <= dias_mes) if tecnico else []
        for d in trabajados:
            fila = dia_a_fila.get(d)
            if fila:
                (escribir_full if patron == "FULL" else escribir_part)(ws, fila)

        etiqueta = hoja if (tecnico is None or tecnico == hoja) else f"{hoja}→{tecnico}"
        resumen.append((etiqueta, patron, len(trabajados), fila_periodo is not None, trabajados))

    wb.save(out_path)

    print("\nHoja             patrón  días  periodo?  (días)")
    for etiqueta, patron, n, tiene_periodo, dd in resumen:
        print(f"  {etiqueta:16} {patron:5} {n:4}   {'sí' if tiene_periodo else 'NO':7} {dd}")
    faltan = {t for t, d in dias.items() if d} - {t for t, _ in MAPEO.values() if t}
    print("\nTécnicos con altas SIN hoja en el registro:", sorted(faltan) or "ninguno")
    print(f"\n✅ Escrito: {out_path}")


if __name__ == "__main__":
    main()
